"""Core orchestration utilities for the AWS security audit."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable, List, Sequence

import boto3

from .findings import Finding, InventoryItem
from .services import SERVICE_CHECKS, ServiceReport


SEVERITY_ORDER = {
    "CRITICAL": 0,
    "ERROR": 1,
    "HIGH": 2,
    "MEDIUM": 3,
    "LOW": 4,
    "WARNING": 5,
    "INFO": 6,
}


def _finding_sort_key(finding: Finding) -> tuple[int, str, str, str]:
    """Return a tuple used to order findings consistently."""

    severity_rank = SEVERITY_ORDER.get(finding.severity.upper(), len(SEVERITY_ORDER))
    return (severity_rank, finding.service, finding.resource_id, finding.message)


@dataclass
class AuditResults:
    """Aggregated findings and inventory from a full audit run."""

    findings: List[Finding]
    inventory: List[InventoryItem]


def collect_audit_results(
    session: boto3.session.Session, services: Iterable[str]
) -> AuditResults:
    """Run all requested service checks and return findings with inventory."""

    findings: dict[str, Finding] = {}
    inventory: List[InventoryItem] = []
    normalized_services: List[str] = []
    for service in services:
        key = service.lower()
        if key not in SERVICE_CHECKS:
            valid = ", ".join(sorted(SERVICE_CHECKS))
            raise ValueError(f"Unknown service '{service}'. Valid services: {valid}")
        normalized_services.append(key)

    for service in dict.fromkeys(normalized_services):
        checker = SERVICE_CHECKS[service]
        report: ServiceReport = checker(session)
        for finding in report.findings:
            findings[finding.key()] = finding
        inventory.extend(report.inventory)

    ordered_findings = sorted(findings.values(), key=_finding_sort_key)
    return AuditResults(findings=ordered_findings, inventory=inventory)


def collect_findings(session: boto3.session.Session, services: Iterable[str]) -> List[Finding]:
    """Run all requested service checks and return de-duplicated findings."""

    return collect_audit_results(session, services).findings


def print_findings(findings: Iterable[Finding]) -> None:
    """Pretty-print findings to stdout."""

    findings = list(findings)
    if not findings:
        print("No findings detected.")
        return

    header = f"{'Service':<10} {'Severity':<8} {'Resource':<40} Message"
    print(header)
    print("-" * len(header))
    for finding in findings:
        resource = (finding.resource_id[:37] + "...") if len(finding.resource_id) > 40 else finding.resource_id
        print(f"{finding.service:<10} {finding.severity:<8} {resource:<40} {finding.message}")


def export_findings_to_excel(findings: Iterable[Finding], path: str) -> str:
    """Write *findings* to an Excel workbook located at *path*."""

    headers = ("Service", "Resource ID", "Severity", "Message")
    rows = (
        (finding.service, finding.resource_id, finding.severity, finding.message)
        for finding in findings
    )
    return _export_rows_to_excel(
        rows,
        headers,
        path,
        sheet_title="Findings",
        purpose="findings",
    )


def export_inventory_to_excel(inventory: Iterable[InventoryItem], path: str) -> str:
    """Write *inventory* to an Excel workbook located at *path*."""

    headers = ("Service", "Resource ID", "Status", "Details")
    rows = (
        (item.service, item.resource_id, item.status, item.details)
        for item in inventory
    )
    return _export_rows_to_excel(
        rows,
        headers,
        path,
        sheet_title="Inventory",
        purpose="inventory",
    )


def _export_rows_to_excel(
    rows: Iterable[Sequence[object]],
    headers: Sequence[str],
    path: str,
    *,
    sheet_title: str,
    purpose: str,
) -> str:
    """Write ``rows`` with ``headers`` to an Excel sheet using :mod:`openpyxl`."""

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - dependency missing during tests
        raise RuntimeError(
            "The 'openpyxl' package is required to export "
            f"{purpose} to Excel. Install it with 'pip install openpyxl'."
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    sheet.append(list(headers))
    column_widths = [len(header) for header in headers]

    for row in rows:
        values = list(row)
        sheet.append(values)
        for idx, value in enumerate(values):
            column_widths[idx] = max(column_widths[idx], len(str(value)))

    for idx, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(idx)
        sheet.column_dimensions[column_letter].width = min(width + 2, 60)

    workbook.save(path)
    return path


__all__ = [
    "AuditResults",
    "collect_audit_results",
    "collect_findings",
    "export_findings_to_excel",
    "export_inventory_to_excel",
    "print_findings",
]
